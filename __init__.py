# 必要なライブラリのインポート
import pandas as pd
import io
import os
from pyspark.sql.window import Window
from openpyxl.styles import Font, Alignment, Border, Side
from pyspark.sql.types import StructType, StructField, StringType, BinaryType
from pyspark.sql import functions as F
from pyspark.sql import DataFrame
from pathlib import Path
import sys
import yaml

from ytcdbutil import DBFSPath
from google_drive import GDFileSystem

from typing import TYPE_CHECKING
from pyspark.dbutils import DBUtils
from pyspark.sql import SparkSession
from pyspark.sql import types as T

def create_output_tariff_navi(config_path, gd, target_month, branch_code) -> DataFrame:
    """
      タリフナビをExcel形式で作成し、google driveの所定の場所に格納するための関数

      Args:
        config_path: 認証情報、drive id等を保存しているyamlファイル
        gd: driveの接続情報
        target_month: 対象年月
        branch_code: 対象主管

      Return:
        result_df: 実行結果
    """
    spark = SparkSession.builder.getOrCreate()

    if TYPE_CHECKING:
        # mypyによる型チェック時のみ実行されるブロック
        spark: SparkSession = SparkSession.builder.getOrCreate()
        dbutils: DBUtils = DBUtils(spark)
    
    with open(config_path) as _f:
        config = yaml.load(_f, Loader=yaml.FullLoader)

    # --- Google Drive 連携に必要な情報を設定 ---
    # 1. Databricks上のプロジェクトのベースディレクトリ
    base_dir = DBFSPath("/dbfs/mnt/datalake003/sandbox/data-agile-group/プロジェクト/タリフナビ")
    # 2. 認証トークンファイルのパス
    token_data = DBFSPath(config["base_dir"] +"/"+ config["google_drive_api"]["token_data"])
    # 3. Databricksシークレットのスコープ名
    secret_scope = config["google_drive_api"]["secret_scope"]
    # 4. Databricksシークレットのキー名
    secret_key = config["google_drive_api"]["secret_key"]

    # 例：「共有ドライブ/チーム共有/プロジェクト/タリフナビ/成果物/」など、対象フォルダのURLからIDをコピーしてください
    target_folder_id = config["google_drive_api"]["drive_path"]

    dm_base_dir = DBFSPath(
      "dbfs:/mnt/datalake003/sandbox/data-agile-group/プロジェクト/タリフナビ/tariff_datamart"
    )
    datamart_dir = dm_base_dir / target_month / branch_code

    customer_master_base_dir = DBFSPath(
      "dbfs:/mnt/datalake003/sandbox/data-agile-group/プロジェクト/タリフナビ/customer_master"
    )
    customer_master_dir = customer_master_base_dir / target_month

    gemini_base_dir = DBFSPath(
      "dbfs:/mnt/datalake003/sandbox/data-agile-group/プロジェクト/タリフナビ/result"
    )
    gemini_dir = gemini_base_dir / branch_code / target_month / "target_gemini_results"

    failed_base_dir = DBFSPath(
      "dbfs:/mnt/datalake003/sandbox/data-agile-group/プロジェクト/タリフナビ/failed_file"
    )
    failed_dir = failed_base_dir / branch_code

    freight_cost_master = DBFSPath(
      "dbfs:/mnt/datalake003/sandbox/data-agile-group/プロジェクト/タリフナビ/freight_cost_master"
    )
    freight_cost_dir = freight_cost_master / target_month

    # Google Driveへのアップロードに成功した際にIDを保存するblobパス 
    gdrive_id_base_dir = DBFSPath(
      "dbfs:/mnt/datalake003/sandbox/data-agile-group/プロジェクト/タリフナビ/gdrive_id"
    )
    gdrive_id_dir = gdrive_id_base_dir / target_month / branch_code

    amazon_master = spark.read.parquet(customer_master_dir.as_dbfs()).filter(
        F.col("Amazon顧客コード判別フラグ") == 1
    )

    # データマート取得（Amazonマスタの情報を削除する）
    dm = spark.read.parquet(datamart_dir.as_dbfs()).join(
        amazon_master, on="顧客コード", how="anti"
    )

    # 運賃マスタ取得
    cost_master = spark.read.parquet(freight_cost_dir.as_dbfs())
    cost_df = cost_master.select(
        F.col("原扱店店所コード").alias("運賃_店所コード"),
        F.col("顧客コード"),
        F.col("サイズ"),
        F.col("着地域名"),
        F.col("運賃").cast("float"),
        F.col("クール運賃").cast("float")
    ).withColumn(
        "商品", F.explode(F.array(F.lit("3商品"), F.lit("ドライ"), F.lit("冷蔵"), F.lit("冷凍")))
    )

    try:
      # gemini解析結果取得
      gemini_df = spark.read.parquet(gemini_dir.as_dbfs())
      gemini_df = gemini_df.select("顧客コード", "action_summary")
      gemini_flag = True
    except Exception as e:
      # gemini結果がなかった場合、空のdf作成
      schema = StructType([
        StructField("顧客コード", StringType(), True),
        StructField("action_summary", StringType(), True)
      ])
      gemini_df = spark.createDataFrame([], schema)
      gemini_flag = False

    # gemini結果をjoin
    dm = dm.join(gemini_df, "顧客コード", "left")

    # 運賃マスタをjoin
    target_cost_df = cost_df.join(dm, "顧客コード", "semi")
    dm = target_cost_df.join(dm, ["顧客コード", "サイズ", "着地域名", "商品"], "left_outer")
    
    # Window関数で、運賃マスタ由来の行 (null) に顧客情報を補完する
    window_spec = Window.partitionBy("顧客コード")
    # 補完するカラムのリスト (DM由来の行には必ず存在するはずのカラム)
    cols_to_fill = ["顧客名", "店所コード", "店所名", "統一コード", "漢字商号"]
    # F.last(col, ignorenulls=True) で、nullでない最後の値を取得し、null を埋める
    for col_name in cols_to_fill:
        dm = dm.withColumn(
            col_name, F.last(F.col(col_name), ignorenulls=True).over(window_spec)
        )

    # Excelのレイアウトを定義
    SIZE_ORDER = ['060', '080', '100', '120', '140', '160', '180', '200', 'コンパクト', 'ネコポス']
    REGION_ORDER = ['北海道', '北東北', '南東北', '関東', '信越', '北陸', '中部', '関西', '中国', '四国', '九州', '沖縄']
    PRODUCT_LIST = ['3商品', 'ドライ', '冷蔵', '冷凍']
    VALUE_COLS = ['単価', '個あたり精緻化原価粗利', '個あたりハキダシ粗利', '合計精緻化粗利', '合計ハキダシ粗利', '個数', '収入合計']
    VALUE_COLS_TITLE = ['単価（円）', '個あたり精緻化粗利（円）：発B以降の精緻化原価で計算', '個あたりハキダシ粗利（円）', '合計精緻化粗利（円）：発B以降の精緻化原価で計算', '合計ハキダシ粗利（円）', '個数（個）', '収入合計（円）']

    # 保存先のベースパス
    BASE_PATH = "/mnt/datalake003/sandbox/SANDBOX/s_tokunaga/タリフナビPJ"

    # Pandas UDFの戻り値のスキーマを定義
    result_schema = StructType([
        StructField("file_path", StringType(), False),
        StructField("file_content", BinaryType(), False)
    ])

    def create_formatted_excel(pdf: pd.DataFrame) -> pd.DataFrame:
        """
        顧客ごとのデータから、整形されたExcelファイルのバイナリデータを生成する。
        """
        # 顧客情報を取得
        first_row = pdf.iloc[0]
        store_code = first_row['店所コード']
        store_name = first_row['店所名']
        customer_code = first_row['顧客コード']
        customer_name = first_row['顧客名']
        touitsu_code = first_row['統一コード']
        company_name = first_row['漢字商号']
        gemini_results = first_row['action_summary']

        # 出力ファイルパスを生成
        file_path = f"{BASE_PATH}/{branch_code}/{customer_code}_{customer_name}/{target_month}_{store_code}_{customer_code}_{customer_name}.xlsx"

        # メモリ上でExcelファイルを扱うための準備
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 商品リストに基づいてシートを作成
            for product in PRODUCT_LIST:
                ws = writer.book.create_sheet(title=product)

                # --- 概要情報の準備と書き込み ---
                product_df = pdf[pdf['商品'] == product]
                if not product_df.empty:
                    total_revenue = product_df['収入合計'].sum()
                    total_seichi_profit = product_df['合計精緻化粗利'].sum()
                    total_hakidashi_profit = product_df['合計ハキダシ粗利'].sum()
                    seichi_profit_rate = float(total_seichi_profit) / total_revenue if total_revenue != 0 else 0
                    hakidashi_profit_rate = float(total_hakidashi_profit) / total_revenue if total_revenue != 0 else 0
                else: # データがない場合は全て0
                    total_revenue, total_seichi_profit, total_hakidashi_profit, seichi_profit_rate, hakidashi_profit_rate = 0, 0, 0, 0, 0

                ws['A1'] = 'タリフナビ'
                ws['C1'] = '法人顧客の収入・個数・単価・粗利を見える化し、単価交渉の際のヒントとしてご活用ください。'
                ws['M1'] = '社外秘'

                ws['B3'] = '年月'
                ws['B4'] = f'{target_month[:4]}年{target_month[4:]}月度'
                ws['C3'] = '店所コード(※1)'
                ws['C4'] = store_code
                ws['D3'] = '店所名'
                ws['D4'] = store_name

                ws['B7'] = '法人統一コード'
                ws['B8'] = touitsu_code
                ws['C7'] = '企業名'
                ws['C8'] = company_name
                ws['D7'] = '顧客コード'
                ws['D8'] = customer_code
                ws['E7'] = '顧客名'
                ws['E8'] = customer_name

                ws['J7'] = '収入合計'
                ws['J8'] = total_revenue
                ws['K7'] = '精緻化粗利合計'
                ws['K8'] = total_seichi_profit
                ws['L7'] = '精緻化粗利率'
                ws['L8'] = seichi_profit_rate
                ws['M7'] = 'ハキダシ粗利合計'
                ws['M8'] = total_hakidashi_profit
                ws['N7'] = 'ハキダシ粗利率'
                ws['N8'] = hakidashi_profit_rate

                ws['C5'] = '(※1)台帳登録店、ただし製販分離適用顧客コードの場合は営業店'
                ws['J3'] = '(※2)精緻化原価とは'
                ws['J4'] = '経営情報の機能別原価をベースにし、不在原価なども考慮し再計算した原価のこと'
                ws['J5'] = 'ハキダシよりも実態に近い、より精緻な原価を目指して算出したもの'
                ws['F23'] = '精緻化原価(※2)ベースで算出した荷物1個あたりの粗利'
                ws['F36'] = 'ハキダシベースで算出した荷物1個あたりの粗利'
                ws['F49'] = '精緻化原価ベースで算出した粗利の合計'
                ws['F62'] = 'ハキダシベースで算出した粗利の合計'

                ws['J1'].value = 'ご意見・お問い合わせはこちら'
                ws['J1'].hyperlink = 'https://docs.google.com/forms/d/e/1FAIpQLSeOuceu1itp8V4dIyvYsWstXE90UhdwsVelPY0e3A9Zc_DW-w/viewform'

                current_row = 10 # 概要情報の下、1行空けて開始

                # --- ピボットテーブルの作成と書き込み ---
                for value_col, title in zip(VALUE_COLS, VALUE_COLS_TITLE):
                    # ピボットテーブルのタイトル
                    if value_col == '単価':
                        if product in ['3商品', 'ドライ']:
                            title = '単価（円）　※顧客台帳の登録運賃を表示しています。'
                        else:
                            title = '単価（円）　※クール料金を含めた単価を表示しています。'
                    ws.cell(row=current_row, column=2, value=title)
                    current_row += 1

                    # 単価テーブルの場合
                    if value_col == '単価':
                        # 単価テーブル用の値を決定
                        if product in ['3商品', 'ドライ']:
                            cost = '運賃'
                        else:
                            cost = 'クール運賃'
                        if not product_df.empty:
                            pivot = pd.pivot_table(product_df, values=cost, index='サイズ', columns='着地域名', aggfunc='sum', fill_value=0)
                            pivot = pivot.reindex(index=SIZE_ORDER, columns=REGION_ORDER, fill_value=0)
                        else: # データがない場合は0埋めのDataFrameを作成
                            pivot = pd.DataFrame(0, index=SIZE_ORDER, columns=REGION_ORDER)

                    else:
                        if not product_df.empty:
                            pivot = pd.pivot_table(product_df, values=value_col, index='サイズ', columns='着地域名', aggfunc='sum', fill_value=0)
                            pivot = pivot.reindex(index=SIZE_ORDER, columns=REGION_ORDER, fill_value=0) # 行と列の順序を固定
                        else: # データがない場合は0埋めのDataFrameを作成
                            pivot = pd.DataFrame(0, index=SIZE_ORDER, columns=REGION_ORDER)

                    # DataFrameをExcelに書き込み (ヘッダーとインデックスを含む)
                    pivot.to_excel(writer, sheet_name=product, startrow=current_row -1, startcol=1) # startrowは1-based
                    current_row += len(pivot) + 2 # ピボットテーブルの下、1行空ける

                # --- 全体のフォーマット調整 ---
                for row in ws.rows:
                    for cell in row:
                        cell.font = Font(name="Arial", size=10)

                for cell in ["B10", "B23", "B36", "B49", "B62", "B75", "B88"]:
                    ws[cell].font = Font(bold=True)

                ws['A1'].font = Font(size=20, name="Arial", bold=True)
                ws['A1'].alignment = Alignment(vertical='center')
                ws['C1'].font = Font(size=12, name="Arial")
                ws['C1'].alignment = Alignment(vertical='center')
                ws['M1'].font = Font(size=18, name="Arial", bold=True, color='FF0000')
                ws['M1'].alignment = Alignment(vertical='center', horizontal='center')
                ws['M1'].border = Border(
                    left=Side(style="thick", color="FF0000"),
                    right=Side(style="thick", color="FF0000"),
                    top=Side(style="thick", color="FF0000"),
                    bottom=Side(style="thick", color="FF0000")
                )

                ws.column_dimensions["A"].width = 5
                ws.row_dimensions[1].height = 42
                ws.column_dimensions["B"].width = 22
                for col_letter in "CDEFGHIJKLMN":
                    ws.column_dimensions[col_letter].width = 16

                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
                border_ranges = ["B3:D4", "B7:E8", "J7:N8", "B11:N21", "B24:N34", "B37:N47", "B50:N60", "B63:N73", "B76:N86", "B89:N99"]
                for range_str in border_ranges:
                    for row in ws[range_str]:
                        for cell in row:
                            cell.border = thin_border

                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                          cell.number_format = '#,##0'
                        if isinstance(cell.value, (int, float)) and cell.value < 0:
                          cell.font = Font(color='FF0000')

                for cell in ["J8", "K8", "M8"]:
                    ws[cell].number_format = "#,##0"
                for cell in ["L8", "N8"]:
                    ws[cell].number_format = "0.00%"

            # Gemini分析結果シートの書き込み
            if gemini_flag and gemini_results is not None:
                ws = writer.book.create_sheet(title="Gemini_分析データ")
                title_font = Font(bold=True, size=12, name="Arial")
                current_row = 2
                # セクションタイトル
                ws.cell(row=current_row, column=2, value="Gemini分析結果").font = title_font
                current_row += 1

                # gemini結果を書き込み
                gemini_cell = ws.cell(row=current_row, column=2, value=gemini_results)
                gemini_cell.font = Font(size=10, name="Arial")
                gemini_cell.alignment = Alignment(wrapText=True)

                # 列幅調整
                ws.column_dimensions["B"].width = 75

            # 初期作成されるデフォルトシートを削除
            if 'Sheet' in writer.book.sheetnames:
                del writer.book['Sheet']

        # 生成したExcelファイルのバイナリデータを返す
        return pd.DataFrame([(file_path, output.getvalue())], columns=["file_path", "file_content"])
      
    print(" 顧客ごとにExcelファイル生成処理を分散実行します...")

    # 顧客コードでグループ化し、各グループに対してPandas UDFを適用
    excel_results_df = dm.groupBy("顧客コード", "店所コード").applyInPandas(create_formatted_excel, schema=result_schema)

    print(f"Google Drive上に主管フォルダ '{branch_code}' を作成します...")

    try:
        # パラメータで指定された主管コード名のフォルダを作成
        # GD_mkdirは既に存在する場合は何もしないので、安全に実行できます
        gd.GD_mkdir(Path(branch_code))
        print(f"主管フォルダ '{branch_code}' の準備が完了しました。")
    except Exception as e:
        print(f"主管フォルダ '{branch_code}' の作成中にエラーが発生しました: {e}")

    @F.udf(T.StringType())
    def get_customer_folder_name(file_path: str) -> str:
        """ file_pathから親ディレクトリ名（顧客フォルダ名）を取得する """
        # file_path例: .../{branch_code}/{customer_code}_{customer_name}/{file_name}.xlsx
        # 欲しいのは `{customer_code}_{customer_name}` の部分
        customer_folder_raw = Path(file_path).parent.name
        # フォルダ名に "/" が含まれるとエラーになるため置換
        customer_code, customer_name = customer_folder_raw.split("_", 1)
        safe_customer_name = customer_name.replace("/", "_")
        return f"{customer_code}_{safe_customer_name}"

    @F.udf(T.StringType())
    def get_upload_file_name(file_path: str) -> str:
        """ file_pathからファイル名を取得する """
        return Path(file_path).name
      
    excel_upload_df = excel_results_df.withColumn(
        "upload_customer_folder", get_customer_folder_name(F.col("file_path"))
    ).withColumn(
        "upload_file_name", get_upload_file_name(F.col("file_path"))
    ).withColumn(
        "branch_code", F.lit(branch_code) # スクリプト冒頭のパラメータを列として追加
    )

    @F.udf(T.StringType())
    def write_excel_to_gdrive(branch_folder: str, customer_folder: str, file_name: str, file_content: bytes) -> str:
        """
        ワーカーノードからGoogle Driveにファイルをアップロードする関数。
        主管フォルダ配下に顧客フォルダを作成し、ファイルをアップロードします。
        """
        # ワーカーノードの一時ディレクトリにファイルを書き出す
        local_temp_path = Path("/tmp") / file_name
        # Google Drive上の最終的な書き込み先フォルダパスを構築
        gdrive_dest_folder = Path(branch_folder) / customer_folder
        # アップロードに失敗したファイルのblob格納先のパスを構築
        failed_upload_folder = failed_dir / customer_folder
        failed_upload_path = failed_upload_folder / file_name

        try:
            # 1. 宛先の顧客フォルダを（必要であれば）作成する
            # この処理は複数のワーカーから同時に呼ばれる可能性がありますが、
            # GD_mkdirがフォルダの存在チェックを行うため、重複作成されることはありません。
            gd.GD_mkdir(gdrive_dest_folder)

            # 2. 一時ファイルにバイナリコンテンツを書き込む
            with open(local_temp_path, "wb") as f:
                f.write(file_content)

            # 3. Google Driveの指定フォルダに、一時ファイルをアップロードする
            # mode='overwrite'で、同名ファイルがあれば上書きします
            result = gd.GD_upload(gdrive_dest_folder, local_temp_path.as_posix(), mode='overwrite')
            id = result.get('id')

            return id if result is not None else 'false'

        except Exception as e:
            # アップロードに失敗した場合、blobフォルダにバイナリコンテンツを書き込む
            print(f"Error uploading {file_name} to {gdrive_dest_folder}: {e}")
            os.makedirs(failed_upload_folder, exist_ok=True)
            with open(failed_upload_path.as_posix(), "wb") as f:
                f.write(file_content)
            return 'false'
        finally:
            # 処理の成否にかかわらず、一時ファイルを必ず削除する
            if local_temp_path.exists():
                local_temp_path.unlink()

    # 並列分散書き込みの実行
    print("生成したExcelファイルをGoogle Driveに並列で書き込みます...")

    result_df = (
        excel_upload_df.repartition(27).withColumn(
            "write_result",
            write_excel_to_gdrive(
                F.col("branch_code"),
                F.col("upload_customer_folder"),
                F.col("upload_file_name"),
                F.col("file_content"),
            ),
        )
    ).checkpoint()

    # 結果をblobに出力
    result_df.write.mode("overwrite").parquet(gdrive_id_dir.as_dbfs())

    return result_df
